from __future__ import annotations

import ast
import calendar
import os
import operator as op
import re
from dataclasses import dataclass
from datetime import date
import json
from pathlib import Path
from typing import Iterable

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from database import (
    SnapshotMonth,
    build_update_frame_from_database,
    build_workbook_data_from_database,
    create_supabase,
    fetch_recent_audit_logs,
    has_app_password,
    insert_audit_log,
    is_supabase_configured,
    list_snapshot_months,
    snapshot_has_values,
    upsert_snapshot_values,
    validate_app_password,
)

DEFAULT_WORKBOOK = Path("/Users/bytedance/Downloads/Personal Assets.xlsx")
SUMMARY_LABELS = {"总资产", "增长率"}
SAFE_OPERATORS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.USub: op.neg,
    ast.UAdd: op.pos,
}


@dataclass
class WorkbookData:
    asset_history: pd.DataFrame
    latest_snapshot: pd.DataFrame
    total_trend: pd.DataFrame
    demo_mode: bool = False


@dataclass
class MonthColumn:
    column_index: int
    label: str
    snapshot_date: date


def evaluate_formula(expr: str) -> float:
    node = ast.parse(expr, mode="eval")

    def _eval(current: ast.AST) -> float:
        if isinstance(current, ast.Expression):
            return _eval(current.body)
        if isinstance(current, ast.Constant) and isinstance(current.value, (int, float)):
            return float(current.value)
        if isinstance(current, ast.BinOp) and type(current.op) in SAFE_OPERATORS:
            return SAFE_OPERATORS[type(current.op)](_eval(current.left), _eval(current.right))
        if isinstance(current, ast.UnaryOp) and type(current.op) in SAFE_OPERATORS:
            return SAFE_OPERATORS[type(current.op)](_eval(current.operand))
        raise ValueError(f"不支持的公式: {expr}")

    return float(_eval(node))


def get_numeric_value(data_ws, formula_ws, row_idx: int, col_idx: int) -> float | None:
    data_value = data_ws.cell(row_idx, col_idx).value
    if isinstance(data_value, (int, float)):
        return float(data_value)

    formula_value = formula_ws.cell(row_idx, col_idx).value
    if isinstance(formula_value, str) and formula_value.startswith("="):
        expr = formula_value[1:].strip()
        if re.fullmatch(r"[0-9+\-*/(). ]+", expr):
            return evaluate_formula(expr)
    return None


def last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def get_month_columns(ws) -> list[MonthColumn]:
    columns: list[MonthColumn] = []
    current_year: int | None = None
    for col in range(3, ws.max_column + 1):
        current_year, parsed_date = parse_header_date(ws.cell(1, col).value, ws.cell(2, col).value, current_year)
        if parsed_date:
            columns.append(
                MonthColumn(
                    column_index=col,
                    label=parsed_date.strftime("%Y-%m-%d"),
                    snapshot_date=parsed_date,
                )
            )
    return columns


def get_asset_rows(ws) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    current_category: str | None = None
    for row_idx in range(3, ws.max_row + 1):
        category_cell = ws.cell(row_idx, 1).value
        account = ws.cell(row_idx, 2).value
        if category_cell in SUMMARY_LABELS:
            break
        if category_cell:
            current_category = str(category_cell).strip()
        if current_category and account:
            rows.append(
                {
                    "row_index": row_idx,
                    "category": current_category,
                    "account": str(account).strip(),
                }
            )
    return rows


def build_update_frame(path: Path, target_column: int) -> pd.DataFrame:
    wb_data = load_workbook(path, data_only=True)
    wb_formula = load_workbook(path, data_only=False)
    ws = wb_data["资产管理"]
    formula_ws = wb_formula["资产管理"]

    month_columns = get_month_columns(ws)
    prev_column = None
    for month_col in month_columns:
        if month_col.column_index < target_column:
            prev_column = month_col.column_index
        if month_col.column_index == target_column:
            break

    rows: list[dict[str, object]] = []
    for item in get_asset_rows(ws):
        previous_value = get_numeric_value(ws, formula_ws, item["row_index"], prev_column) if prev_column else None
        current_value = get_numeric_value(ws, formula_ws, item["row_index"], target_column)
        rows.append(
            {
                "row_index": item["row_index"],
                "分类": item["category"],
                "账户": item["account"],
                "上月金额": previous_value,
                "本月金额": current_value,
            }
        )
    return pd.DataFrame(rows)


def month_has_values(path: Path, target_column: int) -> bool:
    wb_data = load_workbook(path, data_only=True)
    wb_formula = load_workbook(path, data_only=False)
    ws = wb_data["资产管理"]
    formula_ws = wb_formula["资产管理"]
    for item in get_asset_rows(ws):
        if get_numeric_value(ws, formula_ws, item["row_index"], target_column) is not None:
            return True
    return False


def create_month_column(path: Path, target_date: date) -> str:
    wb = load_workbook(path)
    ws = wb["资产管理"]
    month_columns = get_month_columns(ws)
    for item in month_columns:
        if item.snapshot_date == target_date:
            return item.label

    new_col = ws.max_column + 1
    previous_year = month_columns[-1].snapshot_date.year if month_columns else None
    ws.cell(1, new_col).value = f"{target_date.year}年" if previous_year != target_date.year else None
    ws.cell(2, new_col).value = f"金额（{target_date.month}.{target_date.day}）"

    for item in get_asset_rows(ws):
        ws.cell(item["row_index"], new_col).value = None

    total_row = None
    growth_row = None
    for row_idx in range(3, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        if label == "总资产":
            total_row = row_idx
        if label == "增长率":
            growth_row = row_idx
    if total_row:
        start_col_letter = ws.cell(1, 3).column_letter
        end_col_letter = ws.cell(1, new_col).column_letter
        new_col_letter = ws.cell(1, new_col).column_letter
        prev_col_letter = ws.cell(1, new_col - 1).column_letter
        ws.cell(total_row, new_col).value = f"=SUM({new_col_letter}3:{new_col_letter}{total_row - 1})"
        if growth_row:
            ws.cell(growth_row, new_col).value = f"=({new_col_letter}{total_row}-{prev_col_letter}{total_row})/{prev_col_letter}{total_row}"

    wb.save(path)
    return target_date.strftime("%Y-%m-%d")


def save_month_values(path: Path, target_column: int, values: pd.DataFrame) -> None:
    wb = load_workbook(path)
    ws = wb["资产管理"]
    for _, row in values.iterrows():
        row_index = int(row["row_index"])
        amount = row["本月金额"]
        ws.cell(row_index, target_column).value = None if pd.isna(amount) else float(amount)
    wb.save(path)


def add_one_month(source: date) -> date:
    year = source.year + 1 if source.month == 12 else source.year
    month = 1 if source.month == 12 else source.month + 1
    return date(year, month, last_day_of_month(year, month))


def parse_year(value: object, current_year: int | None) -> int | None:
    if value is None:
        return current_year
    if isinstance(value, int):
        return value
    match = re.search(r"(\d{4})", str(value))
    return int(match.group(1)) if match else current_year


def parse_header_date(year_value: object, amount_label: object, current_year: int | None) -> tuple[int | None, date | None]:
    year = parse_year(year_value, current_year)
    if year is None or amount_label is None:
        return year, None
    match = re.search(r"(\d{1,2})\.(\d{1,2})", str(amount_label))
    if not match:
        return year, None
    month, day = map(int, match.groups())
    return year, date(year, month, day)


def parse_assets_sheet(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    wb_data = load_workbook(path, data_only=True)
    wb_formula = load_workbook(path, data_only=False)
    ws = wb_data["资产管理"]
    formula_ws = wb_formula["资产管理"]

    column_dates: dict[int, date] = {}
    current_year: int | None = None
    for col in range(3, ws.max_column + 1):
        current_year, parsed_date = parse_header_date(ws.cell(1, col).value, ws.cell(2, col).value, current_year)
        if parsed_date:
            column_dates[col] = parsed_date

    rows: list[dict[str, object]] = []
    current_category: str | None = None
    for row_idx in range(3, ws.max_row + 1):
        category_cell = ws.cell(row_idx, 1).value
        account = ws.cell(row_idx, 2).value

        if category_cell in SUMMARY_LABELS:
            break
        if category_cell:
            current_category = str(category_cell).strip()
        if not current_category or not account:
            continue

        for col, snapshot_date in column_dates.items():
            amount = get_numeric_value(ws, formula_ws, row_idx, col)
            if amount is not None:
                rows.append(
                    {
                        "category": current_category,
                        "account": str(account).strip(),
                        "date": pd.Timestamp(snapshot_date),
                        "amount": amount,
                    }
                )

    asset_history = pd.DataFrame(rows).sort_values(["date", "category", "account"]).reset_index(drop=True)
    if asset_history.empty:
        raise ValueError("资产管理 工作表中没有可用的数值数据。")

    latest_snapshot = (
        asset_history.sort_values("date")
        .groupby(["category", "account"], as_index=False)
        .tail(1)
        .sort_values(["category", "amount"], ascending=[True, False])
        .reset_index(drop=True)
    )
    return asset_history, latest_snapshot


def build_total_trend(asset_history: pd.DataFrame) -> pd.DataFrame:
    trend = (
        asset_history.groupby("date", as_index=False)["amount"]
        .sum()
        .rename(columns={"amount": "total_assets"})
        .sort_values("date")
        .reset_index(drop=True)
    )
    if trend.empty:
        raise ValueError("资产管理 工作表中没有可用的趋势数据。")
    return trend


def build_demo_workbook_data() -> WorkbookData:
    demo_rows = [
        ("现金", "招行活期", "2025-10-31", 120000),
        ("现金", "招行活期", "2025-11-30", 122000),
        ("现金", "招行活期", "2025-12-31", 125000),
        ("现金", "招行活期", "2026-01-31", 128000),
        ("现金", "招行活期", "2026-02-28", 126000),
        ("现金", "招行活期", "2026-03-31", 130000),
        ("基金", "指数基金A", "2025-10-31", 180000),
        ("基金", "指数基金A", "2025-11-30", 186000),
        ("基金", "指数基金A", "2025-12-31", 191000),
        ("基金", "指数基金A", "2026-01-31", 194000),
        ("基金", "指数基金A", "2026-02-28", 201000),
        ("基金", "指数基金A", "2026-03-31", 208000),
        ("基金", "黄金ETF", "2025-10-31", 50000),
        ("基金", "黄金ETF", "2025-11-30", 52000),
        ("基金", "黄金ETF", "2025-12-31", 51500),
        ("基金", "黄金ETF", "2026-01-31", 54000),
        ("基金", "黄金ETF", "2026-02-28", 56000),
        ("基金", "黄金ETF", "2026-03-31", 57500),
        ("股票", "美股账户", "2025-10-31", 260000),
        ("股票", "美股账户", "2025-11-30", 268000),
        ("股票", "美股账户", "2025-12-31", 275000),
        ("股票", "美股账户", "2026-01-31", 271000),
        ("股票", "美股账户", "2026-02-28", 282000),
        ("股票", "美股账户", "2026-03-31", 296000),
        ("保险", "养老保险账户", "2025-10-31", 90000),
        ("保险", "养老保险账户", "2025-11-30", 90500),
        ("保险", "养老保险账户", "2025-12-31", 91000),
        ("保险", "养老保险账户", "2026-01-31", 91800),
        ("保险", "养老保险账户", "2026-02-28", 92500),
        ("保险", "养老保险账户", "2026-03-31", 93200),
    ]
    asset_history = pd.DataFrame(demo_rows, columns=["category", "account", "date", "amount"])
    asset_history["date"] = pd.to_datetime(asset_history["date"])
    asset_history = asset_history.sort_values(["date", "category", "account"]).reset_index(drop=True)
    latest_snapshot = (
        asset_history.sort_values("date")
        .groupby(["category", "account"], as_index=False)
        .tail(1)
        .sort_values(["category", "amount"], ascending=[True, False])
        .reset_index(drop=True)
    )
    total_trend = build_total_trend(asset_history)
    return WorkbookData(
        asset_history=asset_history,
        latest_snapshot=latest_snapshot,
        total_trend=total_trend,
        demo_mode=True,
    )


@st.cache_data(show_spinner=False)
def load_data(path_str: str) -> WorkbookData:
    if is_supabase_configured():
        client = create_supabase()
        asset_history, latest_snapshot, total_trend = build_workbook_data_from_database(client)
        return WorkbookData(
            asset_history=asset_history,
            latest_snapshot=latest_snapshot,
            total_trend=total_trend,
        )

    path = Path(path_str).expanduser()
    if not path.exists():
        return build_demo_workbook_data()

    asset_history, latest_snapshot = parse_assets_sheet(path)
    total_trend = build_total_trend(asset_history)
    return WorkbookData(
        asset_history=asset_history,
        latest_snapshot=latest_snapshot,
        total_trend=total_trend,
    )


def format_money(value: float) -> str:
    return f"¥{value:,.0f}"


def format_money_compact(value: float) -> str:
    if abs(value) >= 10000:
        return f"¥{value / 10000:.1f}万"
    return format_money(value)


def format_audit_action(action: object) -> str:
    action_map = {
        "save_month_values": "保存月度数据",
    }
    if action is None or (isinstance(action, float) and pd.isna(action)):
        return ""
    return action_map.get(str(action), str(action))


def format_audit_created_at(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    parsed = pd.to_datetime(value, utc=True, errors="coerce")
    if pd.isna(parsed):
        return str(value)
    return parsed.tz_convert("Asia/Shanghai").strftime("%Y-%m-%d %H:%M")


def format_audit_details(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    details = value
    if isinstance(value, str):
        try:
            details = json.loads(value)
        except json.JSONDecodeError:
            return value

    if isinstance(details, dict):
        mode = details.get("mode")
        filled_rows = details.get("filled_rows")
        parts: list[str] = []
        if mode:
            parts.append(str(mode))
        if filled_rows is not None:
            parts.append(f"填写 {filled_rows} 项")
        if parts:
            return "，".join(parts)
        return json.dumps(details, ensure_ascii=False)

    if isinstance(details, list):
        return "；".join(str(item) for item in details)

    return str(details)


def safe_pct_change(values: Iterable[float]) -> float | None:
    values = list(values)
    if len(values) < 2 or values[-2] == 0:
        return None
    return (values[-1] - values[-2]) / values[-2]


def inject_responsive_styles() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.5rem;
            padding-bottom: 2rem;
            max-width: 1240px;
        }

        .asset-kpi-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 0.9rem;
            margin: 0.5rem 0 1rem 0;
        }

        .asset-kpi-card {
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 18px;
            padding: 0.9rem 1rem 0.85rem;
            background: linear-gradient(180deg, rgba(248, 250, 252, 0.95), rgba(255, 255, 255, 1));
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04);
        }

        .asset-section-caption {
            color: #667085;
            font-size: 0.88rem;
            margin: 0 0 0.4rem 0;
        }

        .asset-summary-list {
            display: flex;
            flex-direction: column;
            gap: 0.55rem;
        }

        .asset-summary-row {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 0.9rem;
            padding: 0.72rem 0.85rem;
            border-radius: 14px;
            background: #ffffff;
            border: 1px solid rgba(15, 23, 42, 0.08);
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.03);
        }

        .asset-summary-name {
            font-size: 0.88rem;
            color: #475467;
            flex: 1 1 auto;
            min-width: 0;
        }

        .asset-summary-value {
            font-size: 0.92rem;
            font-weight: 700;
            color: #111827;
            text-align: right;
            white-space: nowrap;
        }

        div[data-testid="stExpander"] {
            margin-top: 2rem !important;
        }

        .asset-kpi-label {
            font-size: 0.86rem;
            color: #667085;
            margin-bottom: 0.45rem;
        }

        .asset-kpi-value {
            font-size: 1.55rem;
            line-height: 1.15;
            font-weight: 700;
            color: #111827;
            letter-spacing: -0.02em;
        }

        .asset-kpi-delta {
            margin-top: 0.5rem;
            display: inline-flex;
            align-items: center;
            gap: 0.3rem;
            border-radius: 999px;
            padding: 0.22rem 0.6rem;
            font-size: 0.84rem;
            font-weight: 600;
            color: #15803d;
            background: #e8f8ee;
        }

        @media (max-width: 768px) {
            .block-container {
                padding-top: 3rem;
                padding-left: 0.7rem;
                padding-right: 0.7rem;
                padding-bottom: 1rem;
                max-width: 100%;
            }

            h1 {
                font-size: 1.45rem !important;
                line-height: 1.2 !important;
                margin-top: 0.35rem !important;
                margin-bottom: 0.35rem !important;
            }

            h3 {
                font-size: 1.02rem !important;
                margin-top: 0.4rem !important;
                margin-bottom: 0.35rem !important;
            }

            p, label, div {
                -webkit-font-smoothing: antialiased;
            }

            .asset-kpi-grid {
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 0.55rem;
                margin: 0.35rem 0 0.8rem 0;
            }

            .asset-kpi-card {
                border-radius: 14px;
                padding: 0.72rem 0.78rem 0.7rem;
                box-shadow: 0 4px 14px rgba(15, 23, 42, 0.04);
            }

            .asset-kpi-label {
                font-size: 0.76rem;
                margin-bottom: 0.3rem;
            }

            .asset-kpi-value {
                font-size: 1.08rem;
                line-height: 1.18;
            }

            .asset-kpi-delta {
                font-size: 0.74rem;
                margin-top: 0.35rem;
                padding: 0.16rem 0.48rem;
            }

            .asset-section-caption {
                font-size: 0.78rem;
                margin-bottom: 0.25rem;
            }

            .asset-summary-list {
                gap: 0.4rem;
            }

            .asset-summary-row {
                border-radius: 12px;
                padding: 0.58rem 0.68rem;
                gap: 0.55rem;
            }

            .asset-summary-name {
                font-size: 0.8rem;
            }

            .asset-summary-value {
                font-size: 0.82rem;
            }

            div[data-testid="stDataFrame"] {
                font-size: 0.82rem !important;
            }

            div[data-testid="stMarkdownContainer"] p {
                font-size: 0.88rem;
                line-height: 1.38;
            }

            div[data-testid="stCaptionContainer"] {
                font-size: 0.78rem !important;
            }

            div[data-testid="stHorizontalBlock"] {
                gap: 0.5rem !important;
            }

            div[data-testid="column"] {
                min-width: 0 !important;
            }

            div[data-testid="stVerticalBlock"] > div:has(> div[data-testid="stMetric"]) {
                gap: 0.4rem !important;
            }

            div[data-baseweb="select"] > div {
                min-height: 44px;
                font-size: 0.92rem;
            }

            button[kind],
            div[data-testid="stButton"] button {
                min-height: 44px;
                border-radius: 12px !important;
                font-size: 0.92rem !important;
            }

            div[data-testid="stExpander"] details {
                border-radius: 14px !important;
            }

            div[data-testid="stDataFrame"] [role="table"] {
                font-size: 0.8rem !important;
            }

            div[role="radiogroup"] {
                gap: 0.3rem !important;
                flex-wrap: wrap !important;
            }

            div[role="radiogroup"] label {
                min-height: 36px !important;
                padding: 0.05rem 0.15rem !important;
            }

            iframe {
                border-radius: 14px !important;
            }

            div[data-testid="stExpander"] {
                margin-top: 1.25rem !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_kpi_grid(items: list[dict[str, str]]) -> None:
    chunks: list[str] = ['<div class="asset-kpi-grid">']
    for item in items:
        chunks.extend(
            [
                '<div class="asset-kpi-card">',
                f'<div class="asset-kpi-label">{item["label"]}</div>',
                f'<div class="asset-kpi-value">{item["value"]}</div>',
                f'<div class="asset-kpi-delta">{item["delta"]}</div>',
                "</div>",
            ]
        )
    chunks.append("</div>")
    st.markdown("".join(chunks), unsafe_allow_html=True)


def render_summary_list(rows: list[tuple[str, str]]) -> None:
    chunks: list[str] = ['<div class="asset-summary-list">']
    for name, value in rows:
        chunks.extend(
            [
                '<div class="asset-summary-row">',
                f'<div class="asset-summary-name">{name}</div>',
                f'<div class="asset-summary-value">{value}</div>',
                "</div>",
            ]
        )
    chunks.append("</div>")
    st.markdown("".join(chunks), unsafe_allow_html=True)


def require_login() -> None:
    if not has_app_password():
        return
    if st.session_state.get("asset_app_authenticated"):
        return

    st.title("资产看板登录")
    st.caption("这是一个受保护的资产应用，请输入访问密码。")
    password = st.text_input("访问密码", type="password")
    if st.button("进入应用", type="primary", use_container_width=True):
        if validate_app_password(password):
            st.session_state["asset_app_authenticated"] = True
            st.rerun()
        st.error("密码不正确，请重试。")
    st.stop()


def render_logout_button() -> None:
    if has_app_password() and st.session_state.get("asset_app_authenticated"):
        if st.button("退出登录", use_container_width=True):
            st.session_state.pop("asset_app_authenticated", None)
            st.rerun()


def build_monthly_analysis(total_trend: pd.DataFrame, asset_history: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    monthly = total_trend.copy()
    monthly["delta_amount"] = monthly["total_assets"].diff()
    monthly["delta_pct"] = monthly["total_assets"].pct_change()

    if len(total_trend) < 2:
        return monthly, pd.DataFrame(columns=["category", "delta_amount"])

    latest_date = total_trend["date"].iloc[-1]
    previous_date = total_trend["date"].iloc[-2]
    latest_category = (
        asset_history[asset_history["date"] == latest_date]
        .groupby("category", as_index=False)["amount"]
        .sum()
        .rename(columns={"amount": "latest_amount"})
    )
    previous_category = (
        asset_history[asset_history["date"] == previous_date]
        .groupby("category", as_index=False)["amount"]
        .sum()
        .rename(columns={"amount": "previous_amount"})
    )
    contribution = latest_category.merge(previous_category, on="category", how="outer").fillna(0)
    contribution["delta_amount"] = contribution["latest_amount"] - contribution["previous_amount"]
    contribution = contribution.sort_values("delta_amount", ascending=False).reset_index(drop=True)
    return monthly, contribution


def build_anomaly_messages(update_frame: pd.DataFrame, monthly_analysis: pd.DataFrame) -> list[str]:
    messages: list[str] = []
    if not update_frame.empty:
        missing_count = int(update_frame["本月金额"].isna().sum())
        if missing_count:
            messages.append(f"当前月份还有 {missing_count} 个账户未填写金额。")
        zero_count = int((update_frame["本月金额"].fillna(0) == 0).sum())
        if zero_count:
            messages.append(f"当前月份有 {zero_count} 个账户金额为 0，请确认是否为真实值。")

    if len(monthly_analysis) >= 2:
        latest = monthly_analysis.iloc[-1]
        if pd.notna(latest["delta_pct"]) and abs(float(latest["delta_pct"])) >= 0.15:
            messages.append(
                f"最近一个月总资产变动 {float(latest['delta_pct']):.2%}，波动较大，建议核对录入和估值。"
            )
    return messages


def dataframe_to_csv_download(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def main() -> None:
    st.set_page_config(page_title="个人资产看板", page_icon="💹", layout="wide")
    inject_responsive_styles()
    require_login()
    prelim_demo_mode = not is_supabase_configured() and not DEFAULT_WORKBOOK.exists()

    with st.sidebar:
        st.header("数据源")
        render_logout_button()
        if is_supabase_configured():
            workbook_path = str(DEFAULT_WORKBOOK)
            refresh_now = st.button("刷新云端数据", use_container_width=True)
            st.caption("当前已接入 Supabase。页面读写都会直接访问云端数据库。")
        else:
            workbook_path = st.text_input("Excel 路径", value=str(DEFAULT_WORKBOOK))
            refresh_now = st.button("重新读取 Excel", use_container_width=True)
            if prelim_demo_mode:
                st.warning("当前未检测到本地 Excel 文件，已自动切换到演示数据模式。")
                st.caption("部署到云端后，如果还没接 Supabase，也会先展示这份演示数据。")
            else:
                st.caption("点一次“重新读取 Excel”就会重新加载第一张工作表。")

    if refresh_now:
        load_data.clear()

    data = load_data(workbook_path)
    use_database = is_supabase_configured()
    demo_mode = data.demo_mode
    if use_database:
        backend_mode = "数据库"
    elif demo_mode:
        backend_mode = "演示"
    else:
        backend_mode = "Excel"

    st.title("个人资产看板")
    st.caption(f"当前数据源：{backend_mode}。自动汇总总资产趋势、分类占比和账户快照。")

    if use_database:
        month_columns = list_snapshot_months(create_supabase())
    elif demo_mode:
        month_columns = [
            SnapshotMonth(
                label=item.strftime("%Y-%m-%d"),
                snapshot_date=item,
                has_values=True,
            )
            for item in data.total_trend["date"].dt.date.tolist()
        ]
    else:
        workbook_file = Path(workbook_path).expanduser()
        wb_headers = load_workbook(workbook_file, data_only=True)
        month_columns = get_month_columns(wb_headers["资产管理"])
    latest_data_date = data.total_trend["date"].iloc[-1].date()
    default_update_date = add_one_month(latest_data_date)
    if use_database:
        display_months = month_columns.copy()
        if not any(item.snapshot_date == default_update_date for item in display_months):
            display_months.append(
                SnapshotMonth(
                    label=default_update_date.strftime("%Y-%m-%d"),
                    snapshot_date=default_update_date,
                    has_values=False,
                )
            )
        display_months = sorted(display_months, key=lambda item: item.snapshot_date)
    elif demo_mode:
        display_months = month_columns
    else:
        display_months = month_columns
    default_update_label = next(
        (item.label for item in display_months if item.snapshot_date == default_update_date),
        display_months[-1].label,
    )
    latest_total = float(data.total_trend["total_assets"].iloc[-1])
    previous_total = float(data.total_trend["total_assets"].iloc[-2]) if len(data.total_trend) > 1 else latest_total
    delta_amount = latest_total - previous_total
    delta_pct = safe_pct_change(data.total_trend["total_assets"].tolist())
    latest_date = data.total_trend["date"].iloc[-1].strftime("%Y-%m-%d")
    monthly_analysis, contribution = build_monthly_analysis(data.total_trend, data.asset_history)

    category_summary = (
        data.latest_snapshot.groupby("category", as_index=False)["amount"].sum().sort_values("amount", ascending=False)
    )
    largest_account = data.latest_snapshot.sort_values("amount", ascending=False).iloc[0]

    render_kpi_grid(
        [
            {
                "label": "最新总资产",
                "value": format_money(latest_total),
                "delta": f"↑ {format_money(delta_amount)}",
            },
            {
                "label": "环比变化",
                "value": f"{delta_pct:.2%}" if delta_pct is not None else "-",
                "delta": f"截至 {latest_date}",
            },
            {
                "label": "跟踪账户数",
                "value": str(data.latest_snapshot["account"].nunique()),
                "delta": "已接入云端数据库" if use_database else "当前为本地模式",
            },
            {
                "label": "最大单项资产",
                "value": str(largest_account["account"]),
                "delta": f"↑ {format_money(float(largest_account['amount']))}",
            },
        ]
    )

    export_left, export_right = st.columns([1.2, 1])
    with export_left:
        st.download_button(
            "导出总资产趋势 CSV",
            data=dataframe_to_csv_download(data.total_trend.assign(date=lambda df: df["date"].dt.strftime("%Y-%m-%d"))),
            file_name="asset_total_trend.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with export_right:
        st.download_button(
            "导出最新账户快照 CSV",
            data=dataframe_to_csv_download(
                data.latest_snapshot.assign(date=lambda df: df["date"].dt.strftime("%Y-%m-%d"))
            ),
            file_name="asset_latest_snapshot.csv",
            mime="text/csv",
            use_container_width=True,
        )

    chart_left, chart_right = st.columns([1.4, 1])

    with chart_left:
        st.subheader("资产趋势")
        filter_left, filter_right = st.columns([1, 1.2])
        with filter_left:
            dimension = st.segmented_control(
                "筛选维度",
                options=["总资产", "分类", "账户"],
                default="总资产",
                selection_mode="single",
            )
        with filter_right:
            if dimension == "总资产":
                selected_value = None
                st.caption("当前展示全部资产的总额变化")
            else:
                dimension_field = "category" if dimension == "分类" else "account"
                available_values = sorted(data.asset_history[dimension_field].unique().tolist())
                selected_value = st.selectbox(f"选择{dimension}", options=available_values)

        if dimension == "总资产":
            trend_view = data.total_trend.copy()
            trend_view["metric"] = "总资产"
            trend_view["amount"] = trend_view["total_assets"]
        else:
            filtered_history = data.asset_history[data.asset_history[dimension_field] == selected_value].copy()
            trend_view = (
                filtered_history.groupby("date", as_index=False)["amount"].sum().sort_values("date").reset_index(drop=True)
            )
            trend_view["metric"] = selected_value

        trend_view["label"] = trend_view["amount"].map(format_money_compact)
        label_indices = {
            trend_view["date"].idxmax(),
            trend_view["amount"].idxmax(),
            trend_view["amount"].idxmin(),
        }
        label_view = trend_view.loc[sorted(label_indices)].copy()
        title_name = "总资产" if dimension == "总资产" else selected_value
        trend_base = alt.Chart(trend_view).encode(
            x=alt.X("date:T", title=None, axis=alt.Axis(format="%Y-%m")),
            y=alt.Y("amount:Q", title=None),
            tooltip=[
                alt.Tooltip("date:T", title="日期", format="%Y-%m-%d"),
                alt.Tooltip("metric:N", title="对象"),
                alt.Tooltip("amount:Q", title="金额", format=",.2f"),
                alt.Tooltip("label:N", title="简写"),
            ],
        )
        trend_chart = (
            trend_base.mark_line(point=alt.OverlayMarkDef(size=70), strokeWidth=3)
            + alt.Chart(label_view).mark_text(dy=-14, fontSize=11, color="#1f2937").encode(
                x="date:T",
                y="amount:Q",
                text="label:N",
            )
        ).properties(height=320)
        st.markdown(f'<div class="asset-section-caption">当前展示：{title_name}</div>', unsafe_allow_html=True)
        st.altair_chart(trend_chart, use_container_width=True)

    with chart_right:
        st.subheader("最新资产分类占比")
        category_summary_view = category_summary.copy()
        category_chart = (
            alt.Chart(category_summary_view)
            .mark_bar()
            .encode(
                x=alt.X("category:N", title=None, axis=alt.Axis(labelAngle=0)),
                y=alt.Y("amount:Q", title=None),
                tooltip=[
                    alt.Tooltip("category:N", title="分类"),
                    alt.Tooltip("amount:Q", title="金额", format=",.2f"),
                ],
            )
            .properties(height=320)
        )
        st.altair_chart(category_chart, use_container_width=True)

    summary_left, summary_right = st.columns([1.4, 1])

    with summary_left:
        st.markdown('<div class="asset-section-caption">趋势摘要</div>', unsafe_allow_html=True)
        trend_summary_rows = [
            (item["date"].strftime("%Y-%m-%d"), format_money(float(item["amount"])))
            for _, item in trend_view.tail(6).iterrows()
        ]
        render_summary_list(trend_summary_rows)

    with summary_right:
        st.markdown('<div class="asset-section-caption">分类摘要</div>', unsafe_allow_html=True)
        category_summary_rows = [
            (str(item["category"]), format_money(float(item["amount"])))
            for _, item in category_summary.head(5).iterrows()
        ]
        render_summary_list(category_summary_rows)

    analysis_left, analysis_right = st.columns([1.1, 1])
    with analysis_left:
        st.subheader("收益分析")
        latest_monthly = monthly_analysis.tail(6).copy()
        latest_monthly["date"] = latest_monthly["date"].dt.strftime("%Y-%m-%d")
        latest_monthly["delta_amount"] = latest_monthly["delta_amount"].apply(
            lambda x: format_money(float(x)) if pd.notna(x) else "-"
        )
        latest_monthly["delta_pct"] = latest_monthly["delta_pct"].apply(
            lambda x: f"{float(x):.2%}" if pd.notna(x) else "-"
        )
        st.dataframe(
            latest_monthly[["date", "delta_amount", "delta_pct"]],
            use_container_width=True,
            hide_index=True,
            height=260,
        )

    with analysis_right:
        st.subheader("分类贡献")
        contribution_rows = [
            (str(item["category"]), format_money(float(item["delta_amount"])))
            for _, item in contribution.head(5).iterrows()
        ]
        render_summary_list(contribution_rows or [("暂无数据", "-")])

    with st.expander("月度更新入口", expanded=False):
        if demo_mode:
            st.info("当前是演示模式：可以查看图表和导出数据，但不会写入 Excel 或数据库。")
            st.caption("如果要开启真实月度录入，请在云端配置 Supabase，或在本机提供可访问的 Excel 文件。")
            return
        update_left, update_right = st.columns([1.1, 1.4])

        with update_left:
            st.markdown("**1. 选择更新方式**")
            update_mode = st.segmented_control(
                "更新方式",
                options=["新增月份录入", "编辑已有月份"],
                default="新增月份录入",
                selection_mode="single",
            )

            if update_mode == "新增月份录入":
                if use_database:
                    st.caption("用于录入新的月份数据。数据库模式下不需要先建列，直接选月份保存即可。")
                else:
                    st.caption("用于录入下一个月份的数据。若该月份列不存在，可先创建。")
                selected_label = st.selectbox(
                    "录入月份",
                    options=[item.label for item in display_months],
                    index=[item.label for item in display_months].index(default_update_label),
                    key="create_month_label",
                )

                new_month_date = st.date_input("新建月份列", value=default_update_date, format="YYYY-MM-DD")

                if not use_database and st.button("创建这个月份列", use_container_width=True):
                    created_label = create_month_column(workbook_file, new_month_date)
                    load_data.clear()
                    st.success(f"已创建月份列：{created_label}")
                    st.rerun()
                if use_database:
                    st.info("数据库模式下，不需要创建列。直接选择日期并保存即可。")
            else:
                filled_months = [
                    item
                    for item in display_months
                    if (
                        snapshot_has_values(create_supabase(), item.snapshot_date)
                        if use_database
                        else month_has_values(workbook_file, item.column_index)
                    )
                ]
                edit_default_label = next(
                    (item.label for item in filled_months if item.snapshot_date == latest_data_date),
                    filled_months[-1].label,
                )
                st.caption("用于二次修正已经录入过的月份，保存后会覆盖该月原值。")
                selected_label = st.selectbox(
                    "编辑月份",
                    options=[item.label for item in filled_months],
                    index=[item.label for item in filled_months].index(edit_default_label),
                    key="edit_month_label",
                )

        with update_right:
            selected_month = next(item for item in display_months if item.label == selected_label)
            if use_database:
                target_column = None
                has_existing_values = snapshot_has_values(create_supabase(), selected_month.snapshot_date)
            else:
                target_column = next(item.column_index for item in month_columns if item.label == selected_label)
                has_existing_values = month_has_values(workbook_file, target_column)
            if update_mode == "编辑已有月份":
                st.markdown("**2. 编辑已有月份并覆盖保存**")
                st.warning(f"{selected_label} 已有数据，下面可以直接修改并覆盖保存。")
            elif has_existing_values:
                st.markdown("**2. 当前月份已有内容，可继续补充或修正**")
                st.info(f"{selected_label} 已存在部分数据，再次保存会覆盖你在下表改动的项目。")
            else:
                st.markdown("**2. 在下表填写本月金额并保存**")

            if use_database:
                update_frame = build_update_frame_from_database(create_supabase(), selected_month.snapshot_date)
                editor_df = update_frame[["asset_id", "分类", "账户", "上月金额", "本月金额"]].copy()
            else:
                update_frame = build_update_frame(workbook_file, target_column)
                editor_df = update_frame[["row_index", "分类", "账户", "上月金额", "本月金额"]].copy()

            anomaly_messages = build_anomaly_messages(update_frame, monthly_analysis)
            for message in anomaly_messages:
                st.warning(message)

            if st.button("用上月金额填充空白项", use_container_width=True):
                fill_mask = editor_df["本月金额"].isna()
                editor_df.loc[fill_mask, "本月金额"] = editor_df.loc[fill_mask, "上月金额"]

            edited_df = st.data_editor(
                editor_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "row_index": st.column_config.NumberColumn("row_index", disabled=True),
                    "asset_id": st.column_config.NumberColumn("asset_id", disabled=True),
                    "分类": st.column_config.TextColumn("分类", disabled=True),
                    "账户": st.column_config.TextColumn("账户", disabled=True),
                    "上月金额": st.column_config.NumberColumn("上月金额", format="%.2f", disabled=True),
                    "本月金额": st.column_config.NumberColumn("本月金额", format="%.2f"),
                },
                disabled=["row_index", "asset_id", "分类", "账户", "上月金额"],
                key=f"monthly_editor_{selected_label}",
            )

            if has_existing_values:
                save_label = "覆盖保存这个月份"
            else:
                save_label = "保存到数据库" if use_database else "保存本月数据到 Excel"
            if st.button(save_label, type="primary", use_container_width=True):
                if use_database:
                    upsert_snapshot_values(create_supabase(), edited_df, selected_month.snapshot_date)
                    insert_audit_log(
                        create_supabase(),
                        "save_month_values",
                        selected_month.snapshot_date,
                        {
                            "mode": update_mode,
                            "filled_rows": int(edited_df["本月金额"].notna().sum()),
                        },
                    )
                else:
                    save_month_values(workbook_file, target_column, edited_df)
                load_data.clear()
                if has_existing_values:
                    target_name = "数据库" if use_database else "Excel"
                    st.success(f"{selected_label} 已重新保存，修改已覆盖到{target_name}。")
                else:
                    target_name = "数据库" if use_database else "Excel"
                    st.success(f"{selected_label} 已保存到{target_name}。")
                st.rerun()

    if use_database:
        st.subheader("最近操作")
        audit_logs = fetch_recent_audit_logs(create_supabase(), limit=10)
        if audit_logs.empty:
            st.caption("暂时还没有审计记录，执行过月度保存后会显示。")
        else:
            audit_view = audit_logs.copy()
            if "snapshot_date" in audit_view.columns:
                audit_view["snapshot_date"] = audit_view["snapshot_date"].fillna("")
            if "created_at" in audit_view.columns:
                audit_view["created_at"] = audit_view["created_at"].apply(format_audit_created_at)
            if "action" in audit_view.columns:
                audit_view["action"] = audit_view["action"].apply(format_audit_action)
            if "details" in audit_view.columns:
                audit_view["details"] = audit_view["details"].apply(format_audit_details)
            audit_view = audit_view.rename(
                columns={
                    "created_at": "操作时间",
                    "action": "操作类型",
                    "snapshot_date": "对应月份",
                    "details": "说明",
                }
            )
            st.dataframe(
                audit_view[["操作时间", "操作类型", "对应月份", "说明"]],
                use_container_width=True,
                hide_index=True,
                height=260,
            )

    st.subheader("账户最新快照")
    snapshot = data.latest_snapshot.copy()
    snapshot["date"] = snapshot["date"].dt.strftime("%Y-%m-%d")
    snapshot["amount"] = snapshot["amount"].map(format_money)
    st.dataframe(snapshot, use_container_width=True, hide_index=True)

    category_filter = st.multiselect(
        "筛选资产类别",
        options=sorted(data.asset_history["category"].unique().tolist()),
        default=sorted(data.asset_history["category"].unique().tolist()),
    )
    filtered_history = data.asset_history[data.asset_history["category"].isin(category_filter)]
    detail_history = (
        filtered_history.pivot_table(index="date", columns="account", values="amount", aggfunc="last")
        .sort_index()
        .reset_index()
    )
    if not detail_history.empty:
        st.subheader("账户历史明细")
        st.dataframe(
            detail_history.assign(date=lambda df: df["date"].dt.strftime("%Y-%m-%d")),
            use_container_width=True,
            hide_index=True,
        )


if __name__ == "__main__":
    main()
